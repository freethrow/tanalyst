from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseForbidden
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.utils.translation import gettext as _
from django.db import connection
from django.conf import settings
from .models import Activity, Todo
from .forms import ActivityForm
import json
from datetime import datetime as dt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from weasyprint import HTML
import io


def check_edit_permission(activity, user):
    """Check if user can edit the activity, return error response if not."""
    if not activity.can_user_edit(user):
        return HttpResponseForbidden(
            _(
                "You don't have permission to edit this activity. Only responsabili or administrators can make changes."
            )
        )
    return None


def activity_list(request):
    """List all activities with pagination and filtering."""
    # Filter by "my activities" - requires raw MongoDB query
    my_activities_filter = request.GET.get("my_activities")
    if my_activities_filter and request.user.is_authenticated:
        user_id = str(request.user.pk)
        # Use raw MongoDB query via PyMongo for array element matching
        db = connection.database
        mongo_filter = {"responsabili": user_id}

        # Apply other filters to MongoDB query
        year_filter = request.GET.get("anno")
        if year_filter:
            mongo_filter["anno"] = int(year_filter)

        sector_filter = request.GET.get("settore")
        if sector_filter:
            mongo_filter["settore"] = {"$regex": sector_filter, "$options": "i"}

        office_filter = request.GET.get("ufficio")
        if office_filter:
            mongo_filter["ufficio"] = office_filter

        # Get matching activity IDs
        activity_ids = [
            doc["_id"] for doc in db.activities.find(mongo_filter, {"_id": 1})
        ]
        activities = Activity.objects.filter(id__in=activity_ids).order_by("-anno", "-mese", "-data_inizio")
    else:
        activities = Activity.objects.all()

        # Filter by year
        year_filter = request.GET.get("anno")
        if year_filter:
            activities = activities.filter(anno=year_filter)

        # Filter by sector
        sector_filter = request.GET.get("settore")
        if sector_filter:
            activities = activities.filter(settore__icontains=sector_filter)

        # Filter by office (exact match since we have defined choices)
        office_filter = request.GET.get("ufficio")
        if office_filter:
            activities = activities.filter(ufficio=office_filter)

    # Pagination
    paginator = Paginator(activities, 20)  # Show 20 activities per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Get unique years for filter
    years = Activity.objects.values_list("anno", flat=True).distinct().order_by("-anno")

    # Use model's OFFICE_CHOICES for the filter dropdown
    offices = Activity.OFFICE_CHOICES

    context = {
        "page_obj": page_obj,
        "years": years,
        "offices": offices,
        "current_year": request.GET.get("anno"),
        "current_sector": request.GET.get("settore"),
        "current_office": request.GET.get("ufficio"),
        "my_activities": my_activities_filter,
    }
    return render(request, "activities/activity_list.html", context)


def activity_detail(request, slug):
    """View a single activity."""
    activity = get_object_or_404(Activity, slug=slug)

    # Get responsabili users for display
    responsabili_users = activity.get_responsabili_users()

    # Check if current user can edit this activity
    can_edit = activity.can_user_edit(request.user)

    context = {
        "activity": activity,
        "responsabili_users": responsabili_users,
        "can_edit": can_edit,
    }
    return render(request, "activities/activity_detail.html", context)


@login_required
def activity_create(request):
    """Create a new activity."""
    if request.method == "POST":
        form = ActivityForm(request.POST)
        if form.is_valid():
            # Save the activity (slug will be auto-generated)
            activity = form.save(commit=False)

            # Get responsabili user IDs
            responsabili_users = form.cleaned_data.get("responsabili", [])
            responsabili_ids = [
                str(user.pk) for user in responsabili_users if user and user.pk
            ]

            # Ensure current user is included
            user_id = str(request.user.pk)
            if user_id not in responsabili_ids:
                responsabili_ids.append(user_id)

            activity.responsabili = responsabili_ids
            activity.save()

            messages.success(
                request,
                f'Activity "{activity.nome_iniziativa}" created successfully!',
            )
            return redirect("activities:activity_detail", slug=activity.slug)
    else:
        # Pre-select the current user as responsabile
        form = ActivityForm(initial={"responsabili": [request.user]})

    context = {"form": form, "action": "Create"}
    return render(request, "activities/activity_form.html", context)


@login_required
def activity_update(request, slug):
    """Update an existing activity."""
    activity = get_object_or_404(Activity, slug=slug)

    # Check if user has permission to edit
    permission_error = check_edit_permission(activity, request.user)
    if permission_error:
        messages.error(request, _("You don't have permission to edit this activity."))
        return redirect("activities:activity_detail", slug=activity.slug)

    if request.method == "POST":
        form = ActivityForm(request.POST, instance=activity)
        if form.is_valid():
            activity = form.save()
            messages.success(
                request, f'Activity "{activity.nome_iniziativa}" updated successfully!'
            )
            return redirect("activities:activity_detail", slug=activity.slug)
    else:
        form = ActivityForm(instance=activity)

    context = {"form": form, "activity": activity, "action": "Update"}
    return render(request, "activities/activity_form.html", context)


@login_required
def activity_delete(request, slug):
    """Delete an activity."""
    activity = get_object_or_404(Activity, slug=slug)

    # Check if user has permission to delete
    permission_error = check_edit_permission(activity, request.user)
    if permission_error:
        messages.error(request, _("You don't have permission to delete this activity."))
        return redirect("activities:activity_detail", slug=activity.slug)

    if request.method == "POST":
        nome = activity.nome_iniziativa
        activity.delete()
        messages.success(request, f'Activity "{nome}" deleted successfully!')
        return redirect("activities:activity_list")

    context = {"activity": activity}
    return render(request, "activities/activity_confirm_delete.html", context)


@require_POST
def generate_excel_report(request):
    """Generate Excel report for selected activities."""
    try:
        # Parse activity IDs from request
        activity_ids_json = request.POST.get("activity_ids", "[]")
        activity_ids = json.loads(activity_ids_json)

        if not activity_ids:
            messages.error(request, "No activities selected for report generation.")
            return redirect("activities:activity_list")

        # Fetch activities from database
        activities = Activity.objects.filter(id__in=activity_ids).order_by(
            "-anno", "-mese"
        )

        if not activities.exists():
            messages.error(request, "No activities found with the selected IDs.")
            return redirect("activities:activity_list")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activities Report"

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="EB6440", end_color="EB6440", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        cell_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add title
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = f"Attivita' ICE Belgrado Report - Generato iled on {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25

        # Add headers (row 3)
        headers = [
            "Data",
            "Tipo",
            "Iniziativa",
            "Citta",
            "Settore",
            "Ufficio",
            "Description",
        ]
        ws.append([])  # Empty row 2
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Set column widths
        column_widths = [12, 15, 30, 15, 20, 15, 20, 40]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Add activity data
        for activity in activities:
            row_data = [
                f"{activity.mese}/{activity.anno}",
                activity.tipo or "",
                activity.nome_iniziativa or "",
                activity.citta or "",
                activity.settore or "",
                activity.ufficio or "",
                activity.descrizione or "",
            ]
            ws.append(row_data)

            # Apply styling to the row
            row_num = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border

        # Freeze header rows
        ws.freeze_panes = "A4"

        # Set row height for title
        ws.row_dimensions[3].height = 30

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create HTTP response
        response = HttpResponse(
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"activities_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except json.JSONDecodeError:
        messages.error(request, "Invalid activity IDs format.")
        return redirect("activities:activity_list")
    except Exception as e:
        messages.error(request, f"Error generating Excel report: {str(e)}")
        return redirect("activities:activity_list")


@require_POST
def generate_pdf_report(request):
    """Generate PDF report for selected activities using WeasyPrint."""
    try:
        # Parse activity IDs from request
        activity_ids_json = request.POST.get("activity_ids", "[]")
        activity_ids = json.loads(activity_ids_json)

        if not activity_ids:
            messages.error(request, "No activities selected for report generation.")
            return redirect("activities:activity_list")

        # Fetch activities from database
        activities = Activity.objects.filter(id__in=activity_ids).order_by(
            "-anno", "-mese"
        )

        if not activities.exists():
            messages.error(request, "No activities found with the selected IDs.")
            return redirect("activities:activity_list")

        # Get logo path for PDF (use absolute file path for WeasyPrint)
        logo_path = settings.BASE_DIR / "static" / "images" / "ita_logo.png"

        # Calculate total number of persons across all activities
        total_persons = sum(
            a.number_persons for a in activities if a.number_persons is not None
        )

        # Render HTML template
        context = {
            "activities": activities,
            "generated_date": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "total_count": activities.count(),
            "total_persons": total_persons if total_persons > 0 else None,
            "logo_path": logo_path.as_uri() if logo_path.exists() else "",
        }

        html_string = render_to_string("activities/activities_report_pdf.html", context)

        # Generate PDF with base_url for resolving relative paths
        pdf_file = HTML(string=html_string, base_url=str(settings.BASE_DIR)).write_pdf()

        # Create HTTP response
        response = HttpResponse(pdf_file, content_type="application/pdf")
        filename = f"activities_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except json.JSONDecodeError:
        messages.error(request, "Invalid activity IDs format.")
        return redirect("activities:activity_list")
    except Exception as e:
        messages.error(request, f"Error generating PDF report: {str(e)}")
        return redirect("activities:activity_list")


@login_required
@require_POST
def add_todo(request, slug):
    """Add a new todo to an activity via HTMX."""
    activity = get_object_or_404(Activity, slug=slug)

    # Check if user has permission to edit
    if not activity.can_user_edit(request.user):
        return HttpResponse(
            f"<p class=\"text-red-500 text-sm\">{_('You do not have permission to modify this activity.')}</p>",
            status=403,
        )

    name = request.POST.get("name", "").strip()
    description = request.POST.get("description", "").strip()
    datetime_str = request.POST.get("datetime", "").strip()

    if not name:
        return HttpResponse(
            '<p class="text-red-500 text-sm">Name is required</p>', status=400
        )

    # Parse datetime if provided
    todo_datetime = None
    if datetime_str:
        try:
            todo_datetime = dt.fromisoformat(datetime_str)
        except ValueError:
            pass

    # Create new todo
    new_todo = Todo(
        name=name,
        description=description,
        datetime=todo_datetime,
        done=False,
    )

    # Add to activity's todos
    if activity.todos is None:
        activity.todos = []
    activity.todos.append(new_todo)
    activity.save()

    # Return updated todo list partial
    context = {"activity": activity}
    return render(request, "activities/partials/todo_list.html", context)


@login_required
@require_POST
def remove_todo(request, slug, todo_index):
    """Remove a todo from an activity via HTMX."""
    activity = get_object_or_404(Activity, slug=slug)

    # Check if user has permission to edit
    if not activity.can_user_edit(request.user):
        return HttpResponse(
            f"<p class=\"text-red-500 text-sm\">{_('You do not have permission to modify this activity.')}</p>",
            status=403,
        )

    if activity.todos and 0 <= todo_index < len(activity.todos):
        activity.todos.pop(todo_index)
        activity.save()

    # Return updated todo list partial
    context = {"activity": activity}
    return render(request, "activities/partials/todo_list.html", context)


@login_required
@require_POST
def toggle_todo(request, slug, todo_index):
    """Toggle the done status of a todo via HTMX."""
    activity = get_object_or_404(Activity, slug=slug)

    # Check if user has permission to edit
    if not activity.can_user_edit(request.user):
        return HttpResponse(
            f"<p class=\"text-red-500 text-sm\">{_('You do not have permission to modify this activity.')}</p>",
            status=403,
        )

    if activity.todos and 0 <= todo_index < len(activity.todos):
        activity.todos[todo_index].done = not activity.todos[todo_index].done
        activity.save()

    # Return updated todo list partial
    context = {"activity": activity}
    return render(request, "activities/partials/todo_list.html", context)
